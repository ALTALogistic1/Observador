"""Connecteur EIMT positive (Étude d'impact sur le marché du travail / TFWP
Positive LMIA Employers List) — spec section 7, Signal 3.

PROMU en source active Phase 1 (mise à jour de spec du 2026-08-31) pour
remplacer Guichet-Emplois comme source du signal recrutement : contrairement au
Guichet-Emplois, ce jeu de données donne le NOM DE L'EMPLOYEUR — confirmé en
inspectant le vrai fichier (colonnes réelles : Province/Territory, Program
Stream, Employer, Address, Occupation, Incorporate Status, Approved LMIAs,
Approved Positions).

Données ouvertes CKAN (open.canada.ca), jeu de données
90fed587-1364-4f33-a9ee-208181dc0b97, publié trimestriellement.

IMPORTANT — format réel confirmé le 2026-08-31 : les trimestres récents sont en
**XLSX**, pas CSV comme la spec le supposait à l'origine (les plus anciens,
jusqu'à ~2021, sont en CSV) — ce connecteur gère les deux. Le fichier a une ligne
de titre fusionnée avant la vraie ligne d'en-têtes ; `_find_header_row` la
localise en cherchant la colonne "Employer"/"Employeur" plutôt que de supposer un
index de ligne fixe.
"""
from __future__ import annotations

import csv
import logging
import re
from collections.abc import Iterator
from datetime import datetime, timezone

import openpyxl

from observador.sources.base import RawSignal, SourceConnector
from observador.sources.ckan_client import OPEN_CANADA_BASE, CKANClient
from observador.sources.column_mapping import normaliser, resolve_columns

logger = logging.getLogger(__name__)

EIMT_PACKAGE_ID = "90fed587-1364-4f33-a9ee-208181dc0b97"

COLUMN_ALIASES: dict[str, list[str]] = {
    # Alias vérifiés contre les vraies en-têtes EN et FR du fichier 2026Q1
    # ("Province/Territory"/"Province/Territoire", "Program Stream"/"Volet du
    # programme", "Employer"/"Employeur", "Address"/"Adresse",
    # "Occupation"/"Profession", "Approved LMIAs"/"EIMT Approuvés",
    # "Approved Positions"/"Postes Approuvés").
    "employeur": ["employer", "employeur"],
    "adresse": ["address", "adresse"],
    "profession": ["occupation", "profession"],
    "province": ["province"],
    "volet_programme": ["program_stream", "volet"],
    "nombre_postes": ["approved_positions", "postes_approuves"],
    "nombre_lmia": ["approved_lmias", "eimt_approuves"],
}

_RE_TRIMESTRE = re.compile(r"(20\d{2})\s*Q([1-4])", re.IGNORECASE)


def _trimestre_depuis_nom(nom: str) -> datetime:
    """Extrait la date de début du trimestre à partir du nom de la ressource
    (ex. "2026Q1-Employers Who Were..."). Chaque ligne du fichier n'a pas de
    date propre — seul le trimestre de publication en tient lieu (spec :
    "récurrence sur plusieurs trimestres")."""
    m = _RE_TRIMESTRE.search(nom)
    if not m:
        return datetime.now(timezone.utc)
    annee, trimestre = int(m.group(1)), int(m.group(2))
    mois = {1: 1, 2: 4, 3: 7, 4: 10}[trimestre]
    return datetime(annee, mois, 1, tzinfo=timezone.utc)


def _find_header_row(rows: list[tuple]) -> tuple[int, list[str]]:
    """La première ligne du fichier est un TITRE fusionné sur une seule cellule
    (une phrase complète, qui contient elle-même le mot "employeurs" — un simple
    test de sous-chaîne s'y ferait donc piéger). La vraie ligne d'en-têtes est
    identifiée par une cellule dont le contenu normalisé est EXACTEMENT
    "employer"/"employeur" (un nom de colonne court), pas une phrase qui le
    contient."""
    for i, row in enumerate(rows[:5]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any(normaliser(c) in ("employer", "employeur") for c in cells):
            return i, cells
    raise ValueError(
        "Ligne d'en-tête (colonne Employer/Employeur exacte) introuvable dans "
        "les 5 premières lignes du fichier EIMT."
    )


def _read_xlsx(path) -> tuple[dict[str, str], list[dict]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    header_idx, header = _find_header_row(rows)
    columns = resolve_columns(header, COLUMN_ALIASES)
    data_rows = [dict(zip(header, row, strict=False)) for row in rows[header_idx + 1 :]]
    return columns, data_rows


def _read_csv(path) -> tuple[dict[str, str], list[dict]]:
    with open(path, encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.DictReader(f)
        header = list(reader.fieldnames or [])
        columns = resolve_columns(header, COLUMN_ALIASES)
        data_rows = list(reader)
    return columns, data_rows


def _parse_int(raw) -> int | None:
    if raw is None or raw == "":
        return None
    try:
        return int(float(str(raw).replace(",", "").strip()))
    except ValueError:
        return None


class EIMTConnector(SourceConnector):
    def __init__(self, source_def, limit: int | None = None):
        super().__init__(source_def)
        self.limit = limit

    def detect(self, since: datetime | None, db_session) -> Iterator[RawSignal]:
        client = CKANClient(OPEN_CANADA_BASE)
        resources = client.resources(EIMT_PACKAGE_ID, format_filter="XLSX") or client.resources(
            EIMT_PACKAGE_ID, format_filter="CSV"
        )
        if not resources:
            logger.warning("EIMT: aucune ressource XLSX/CSV trouvée sur CKAN")
            return

        # Préférence pour la version française (résultats destinés au Québec en
        # Phase 1) ; repli sur ce qui est disponible si l'attribut "language"
        # n'est pas renseigné pour une ressource donnée.
        fr = [r for r in resources if r.get("language") == ["fr"]]
        resources = fr or resources

        cibles = resources[:1] if since is None else resources[:4]  # ~1 an de trimestres
        lignes_lues = 0

        for resource in cibles:
            path = client.download(resource)
            trimestre_date = _trimestre_depuis_nom(resource.get("name", ""))
            fmt = (resource.get("format") or "").upper()

            try:
                columns, data_rows = _read_xlsx(path) if fmt == "XLSX" else _read_csv(path)
            except ValueError as exc:
                logger.warning("EIMT: impossible de lire %s: %s", resource.get("name"), exc)
                continue

            for i, row in enumerate(data_rows):
                if self.limit is not None and lignes_lues >= self.limit:
                    break

                employeur = str(row.get(columns["employeur"]) or "").strip()
                nb_postes = _parse_int(row.get(columns["nombre_postes"]))
                if not employeur or nb_postes is None:
                    continue  # ligne de note/pied de page plutôt qu'une vraie donnée
                lignes_lues += 1

                if since and trimestre_date < since:
                    continue

                province = str(row.get(columns["province"]) or "").strip() or None
                profession = str(row.get(columns["profession"]) or "").strip() or None

                yield RawSignal(
                    signal_type_id="recrutement_massif",
                    nom_entreprise=employeur,
                    detected_at=trimestre_date,
                    source_ref=f"eimt:{resource['id']}:{i}",
                    region=province,
                    titre_ou_description=profession,
                    valeur_associee=float(nb_postes),
                    champs={
                        "adresse": row.get(columns["adresse"]),
                        "profession": profession,
                        "province": province,
                        "volet_programme": row.get(columns["volet_programme"]),
                        "nombre_postes_approuves": nb_postes,
                        "nombre_lmia_approuvees": _parse_int(row.get(columns["nombre_lmia"])),
                        "trimestre": resource.get("name"),
                    },
                )

            if self.limit is not None and lignes_lues >= self.limit:
                break


CONNECTOR_CLASS = EIMTConnector
